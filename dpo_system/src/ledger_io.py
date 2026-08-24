"""Deterministic ledger reader/writer for operator notebooks.

This module performs strict local spreadsheet I/O only.
It does not trigger ingestion, execution, network calls, or external APIs.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

LEDGER_PATH = Path(__file__).resolve().parents[1] / "operator_ledger.xlsx"

REPO_STATUS = {"OK", "FAIL", "NEEDS_UPDATE"}
INGESTION_TYPES = {"case_law", "contracts", "scholarly", "crm", "custom"}
TASK_STATUS = {"pending", "running", "complete", "failed"}
CLAUSE_ENGINES = {"Atticus", "ClauseX", "HermesLegal"}
ANALYSIS_TYPES = {"classification", "deviation", "scoring", "fallback"}
CRM_ACTIONS = {"email", "campaign", "update", "classify", "enrich"}
GOV_ACTIONS = {"validate", "approve", "reject", "update"}

SHEET_HEADERS: dict[str, list[str]] = {
    "REPO_STATE": [
        "repo_name",
        "repo_slug",
        "wave",
        "enabled",
        "last_pull",
        "last_commit",
        "status",
        "notes",
    ],
    "INGESTION_QUEUE": [
        "batch_id",
        "source_repo",
        "source_path",
        "ingestion_type",
        "priority",
        "status",
        "created_at",
        "updated_at",
        "operator",
        "workflow_id",
        "origin_repo",
        "hop_count",
        "notes",
        "list_name",
        "outreach_purpose",
        "list_source",
        "outbound_status",
    ],
    "CLAUSE_QUEUE": [
        "clause_task_id",
        "contract_id",
        "clause_engine",
        "analysis_type",
        "priority",
        "status",
        "created_at",
        "updated_at",
        "operator",
        "workflow_id",
        "origin_repo",
        "hop_count",
        "notes",
        "list_name",
        "outreach_purpose",
        "list_source",
        "outbound_status",
    ],
    "CRM_QUEUE": [
        "crm_task_id",
        "lead_id",
        "action_type",
        "priority",
        "status",
        "created_at",
        "updated_at",
        "operator",
        "workflow_id",
        "origin_repo",
        "hop_count",
        "notes",
        "list_name",
        "outreach_purpose",
        "list_source",
        "outbound_status",
    ],
    "GOVERNANCE_QUEUE": [
        "governance_id",
        "interview_name",
        "action_type",
        "priority",
        "status",
        "created_at",
        "updated_at",
        "operator",
        "workflow_id",
        "origin_repo",
        "hop_count",
        "notes",
        "list_name",
        "outreach_purpose",
        "list_source",
        "outbound_status",
    ],
    "OPERATOR_LOG": [
        "log_id",
        "timestamp",
        "operator",
        "notebook",
        "action",
        "target",
        "status",
        "notes",
    ],
    "LEDGER_EVENTS": [
        "event_id",
        "timestamp",
        "operator",
        "notebook",
        "action",
        "target",
        "status",
        "notes",
        "item_id",
        "lane",
        "decision",
        "source_id",
        "evidence_ref",
    ],
}

ENUM_FIELDS: dict[str, dict[str, set[str]]] = {
    "REPO_STATE": {
        "status": REPO_STATUS,
    },
    "INGESTION_QUEUE": {
        "ingestion_type": INGESTION_TYPES,
        "status": TASK_STATUS,
    },
    "CLAUSE_QUEUE": {
        "clause_engine": CLAUSE_ENGINES,
        "analysis_type": ANALYSIS_TYPES,
        "status": TASK_STATUS,
    },
    "CRM_QUEUE": {
        "action_type": CRM_ACTIONS,
        "status": TASK_STATUS,
    },
    "GOVERNANCE_QUEUE": {
        "action_type": GOV_ACTIONS,
        "status": TASK_STATUS,
    },
    "OPERATOR_LOG": {
        "status": {"ok", "fail"},
    },
    "LEDGER_EVENTS": {
        "status": {"ok", "fail"},
    },
}

BOOL_FIELDS: dict[str, set[str]] = {
    "REPO_STATE": {"enabled"},
}

INT_FIELDS: dict[str, set[str]] = {
    "REPO_STATE": {"wave"},
    "INGESTION_QUEUE": {"priority", "hop_count"},
    "CLAUSE_QUEUE": {"priority", "hop_count"},
    "CRM_QUEUE": {"priority", "hop_count"},
    "GOVERNANCE_QUEUE": {"priority", "hop_count"},
}

TIMESTAMP_FIELDS: dict[str, set[str]] = {
    "REPO_STATE": {"last_pull"},
    "INGESTION_QUEUE": {"created_at", "updated_at"},
    "CLAUSE_QUEUE": {"created_at", "updated_at"},
    "CRM_QUEUE": {"created_at", "updated_at"},
    "GOVERNANCE_QUEUE": {"created_at", "updated_at"},
    "OPERATOR_LOG": {"timestamp"},
    "LEDGER_EVENTS": {"timestamp"},
}

AUTO_NOW_FIELDS: dict[str, set[str]] = {
    "INGESTION_QUEUE": {"created_at", "updated_at"},
    "CLAUSE_QUEUE": {"created_at", "updated_at"},
    "CRM_QUEUE": {"created_at", "updated_at"},
    "GOVERNANCE_QUEUE": {"created_at", "updated_at"},
    "OPERATOR_LOG": {"timestamp"},
    "LEDGER_EVENTS": {"timestamp"},
}


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def create_ledger_workbook(path: str | Path) -> Path:
    ledger_path = Path(path)
    ledger_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    first_sheet = True
    for sheet_name, headers in SHEET_HEADERS.items():
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)

    wb.save(ledger_path)
    return ledger_path


def timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def validate_enum(allowed_set, value, column_name):
    if value not in allowed_set:
        raise ValueError(
            f"Invalid value '{value}' for column '{column_name}'. Allowed: {allowed_set}"
        )


def _normalize_timestamp(value: Any) -> str:
    if isinstance(value, datetime):
        dt = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            raise ValueError("timestamp string must not be empty")
        if raw.endswith("Z"):
            raw = raw[:-1] + "+00:00"
        try:
            dt = datetime.fromisoformat(raw)
        except ValueError as exc:
            raise ValueError(f"invalid ISO timestamp: {value}") from exc
        dt = dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    raise ValueError(f"unsupported timestamp type: {type(value).__name__}")


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        norm = value.strip().lower()
        if norm in {"true", "1", "yes", "y"}:
            return True
        if norm in {"false", "0", "no", "n"}:
            return False
    if isinstance(value, int) and value in {0, 1}:
        return bool(value)
    raise ValueError(f"cannot coerce to bool: {value}")


def _require_sheet(sheet_name: str) -> None:
    if sheet_name not in SHEET_HEADERS:
        raise ValueError(f"unsupported sheet: {sheet_name}")


def safe_load_workbook():
    if not LEDGER_PATH.exists():
        raise FileNotFoundError(f"Ledger not found at {LEDGER_PATH}")
    return load_workbook(LEDGER_PATH)


def _open_ledger(path: Path | None = None):
    ledger = path or LEDGER_PATH
    if ledger == LEDGER_PATH:
        return safe_load_workbook(), ledger
    if not ledger.exists():
        raise FileNotFoundError(f"ledger file not found: {ledger}")
    return load_workbook(ledger), ledger


def _validate_headers(ws, sheet_name: str) -> None:
    expected = SHEET_HEADERS[sheet_name]
    actual = [cell.value for cell in ws[1]]
    if actual == expected:
        return

    if sheet_name == "GOVERNANCE_QUEUE" and actual == [
        "governance_id",
        "interview_name",
        "action_type",
        "priority",
        "status",
        "created_at",
        "updated_at",
        "operator",
        "notes",
    ]:
        for idx, header in enumerate(expected):
            if idx >= len(actual):
                ws.cell(row=1, column=idx + 1, value=header)
            elif actual[idx] != header:
                ws.cell(row=1, column=idx + 1, value=header)
        return

    raise RuntimeError(
        f"sheet header mismatch for {sheet_name}; expected {expected}, found {actual}"
    )


def list_sheets(path: Path | None = None) -> list[str]:
    wb, _ = _open_ledger(path)
    return list(wb.sheetnames)


def read_rows(sheet_name: str, path: Path | None = None) -> list[dict[str, Any]]:
    _require_sheet(sheet_name)
    wb, _ = _open_ledger(path)
    ws = wb[sheet_name]
    _validate_headers(ws, sheet_name)

    headers = SHEET_HEADERS[sheet_name]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in values):
            continue
        row = {headers[i]: values[i] for i in range(len(headers))}
        rows.append(row)
    return rows


def read_sheet(sheet_name):
    _require_sheet(sheet_name)
    wb = safe_load_workbook()
    ws = wb[sheet_name]
    _validate_headers(ws, sheet_name)
    headers = [c.value for c in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))

    return rows


def _normalize_row_for_write(sheet_name: str, row: dict[str, Any]) -> dict[str, Any]:
    headers = SHEET_HEADERS[sheet_name]
    output: dict[str, Any] = {}

    # Keep deterministic column order and fill missing optional values.
    for key in headers:
        output[key] = row.get(key, "")

    # Enforce primary ID-like first column for every sheet.
    id_col = headers[0]
    if not str(output[id_col]).strip():
        raise ValueError(f"{id_col} is required for {sheet_name}")

    for field in BOOL_FIELDS.get(sheet_name, set()):
        output[field] = _to_bool(output[field])

    for field in INT_FIELDS.get(sheet_name, set()):
        value = output[field]
        if value == "":
            raise ValueError(f"{field} is required for {sheet_name}")
        try:
            output[field] = int(value)
        except Exception as exc:
            raise ValueError(f"{field} must be int for {sheet_name}: {value}") from exc

    # Auto-fill timestamp fields where configured and missing.
    for field in AUTO_NOW_FIELDS.get(sheet_name, set()):
        if output[field] in {"", None}:
            output[field] = _utc_now_iso()

    for field in TIMESTAMP_FIELDS.get(sheet_name, set()):
        value = output[field]
        if value in {"", None}:
            continue
        output[field] = _normalize_timestamp(value)

    enum_rules = ENUM_FIELDS.get(sheet_name, {})
    for field, allowed in enum_rules.items():
        value = output[field]
        if value in {"", None}:
            raise ValueError(f"{field} is required for {sheet_name}")
        value_str = str(value)
        validate_enum(allowed, value_str, f"{sheet_name}.{field}")
        output[field] = value_str

    return output


def append_row(sheet_name, row_dict, path: Path | None = None):
    _require_sheet(sheet_name)
    wb, ledger = _open_ledger(path)
    ws = wb[sheet_name]
    _validate_headers(ws, sheet_name)

    headers = [c.value for c in ws[1]]
    normalized = _normalize_row_for_write(sheet_name, row_dict)
    row = [normalized.get(h, "") for h in headers]

    ws.append(row)
    wb.save(ledger)


def update_row(sheet_name, key_column, key_value, updates, path: Path | None = None):
    _require_sheet(sheet_name)
    wb, ledger = _open_ledger(path)
    ws = wb[sheet_name]
    _validate_headers(ws, sheet_name)

    headers = [c.value for c in ws[1]]
    if key_column not in headers:
        raise ValueError(f"key column '{key_column}' not found in {sheet_name}")

    key_index = headers.index(key_column)
    match_row_index = None

    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[key_index] == key_value:
            match_row_index = row_index
            break

    if match_row_index is None:
        raise KeyError(f"no row found in {sheet_name} where {key_column}={key_value}")

    current = {headers[i]: ws.cell(row=match_row_index, column=i + 1).value for i in range(len(headers))}
    merged = dict(current)
    merged.update(updates)
    normalized = _normalize_row_for_write(sheet_name, merged)

    for i, header in enumerate(headers, start=1):
        ws.cell(row=match_row_index, column=i, value=normalized.get(header, ""))

    wb.save(ledger)
    return normalized


def log_operator_action(operator, notebook, action, target, status="ok", notes="", path: Path | None = None):
    append_row(
        "OPERATOR_LOG",
        {
            "log_id": f"log-{timestamp()}",
            "timestamp": timestamp(),
            "operator": operator,
            "notebook": notebook,
            "action": action,
            "target": target,
            "status": status,
            "notes": notes,
        },
        path=path,
    )


def replace_sheet_rows(
    sheet_name: str, rows: list[dict[str, Any]], path: Path | None = None
) -> list[dict[str, Any]]:
    """Deterministically rewrite a sheet body while preserving schema headers."""
    _require_sheet(sheet_name)
    wb, ledger = _open_ledger(path)
    ws = wb[sheet_name]
    _validate_headers(ws, sheet_name)

    # Clear all data rows only (keep header row untouched).
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    normalized_rows: list[dict[str, Any]] = []
    for row in rows:
        normalized = _normalize_row_for_write(sheet_name, row)
        ws.append([normalized[h] for h in SHEET_HEADERS[sheet_name]])
        normalized_rows.append(normalized)

    wb.save(ledger)
    return normalized_rows
